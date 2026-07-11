"""File reader — loads CSV, Parquet, and Excel files into Polars DataFrames."""

from __future__ import annotations

import logging
from pathlib import Path

from goldencheck._polars_lazy import pl

logger = logging.getLogger(__name__)
SUPPORTED_EXTENSIONS = {".csv", ".parquet", ".xlsx", ".xls"}


def read_file(path: Path) -> pl.DataFrame:
    # Normalise the path early so symlink-and-`..` traversal can't smuggle
    # access to unexpected locations through downstream Polars / openpyxl
    # readers. `path` is a trusted-config / CLI value, but the normalisation
    # is cheap insurance against accidental traversal in API callers.
    path = Path(path).resolve()
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file format: {ext}. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path.name}")

    # Log basename only — full paths can carry control characters /
    # newlines that pollute structured logs (CodeQL py/log-injection).
    logger.info("Reading %s (%s)", path.name, ext)

    if path.stat().st_size == 0:
        raise ValueError("File has no data rows. Nothing to profile.")

    if ext == ".csv":
        try:
            return pl.read_csv(path, infer_schema_length=10000)
        except ImportError:
            # The lazy `pl` proxy's helpful "install goldencheck[polars]" ImportError --
            # let it propagate as-is instead of getting coerced into the generic
            # ValueError below, which would be indistinguishable from a real malformed
            # CSV and would break `except ImportError` callers checking for the
            # missing-Polars case specifically.
            raise
        except Exception:
            try:
                return pl.read_csv(path, infer_schema_length=10000, encoding="latin-1")
            except ImportError:
                raise
            except Exception as e:
                raise ValueError(
                    f"Could not read CSV: {e}. Try specifying --separator or --quote-char"
                ) from e
    elif ext == ".parquet":
        return pl.read_parquet(path)
    elif ext in (".xlsx", ".xls"):
        try:
            return pl.read_excel(path, engine="openpyxl")
        except Exception as e:
            if "password" in str(e).lower() or "encrypted" in str(e).lower():
                raise ValueError(
                    "File appears to be password-protected. GoldenCheck cannot read encrypted files."
                ) from e
            raise
    else:
        raise ValueError(f"Unsupported file format: {ext}")


def _read_csv_columns(path: Path) -> dict[str, list]:
    """Read a CSV file into a dict[str, list] using Polars.

    CSV cannot be read byte-identically without Polars (its dtype inference isn't
    reproducible), so this requires Polars and mirrors `read_file`'s CSV path exactly.
    """
    try:
        import polars  # noqa: F401
    except ImportError as e:
        raise ImportError(
            "Reading CSV requires Polars (its dtype inference is not reproducible without "
            "it). Install goldencheck[polars], or use a Parquet/Excel source."
        ) from e
    try:
        df = pl.read_csv(path, infer_schema_length=10000)
    except Exception:
        df = pl.read_csv(path, infer_schema_length=10000, encoding="latin-1")
    return df.to_dict(as_series=False)


def read_columns(path: Path) -> dict[str, list]:
    """Polars-free typed read into columns for scan_columns(). Parquet (pyarrow) and
    Excel (openpyxl) read without Polars; CSV needs Polars (dtype inference). Returns
    {column_name: [values...]} with native Python scalars matching the pl.read_* values."""
    path = Path(path).resolve()
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file format: {ext}. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path.name}")
    logger.info("Reading %s (%s) [columns]", path.name, ext)
    if path.stat().st_size == 0:
        raise ValueError("File has no data rows. Nothing to profile.")
    if ext == ".parquet":
        return _read_parquet_columns(path)
    if ext in (".xlsx", ".xls"):
        return _read_excel_columns(path)
    if ext == ".csv":
        return _read_csv_columns(path)
    raise ValueError(f"Unsupported file format: {ext}")


def _read_parquet_columns(path: Path) -> dict[str, list]:
    """Read a Parquet file into a dict[str, list] without Polars.

    Byte-identical to `pl.read_parquet(path).to_dict(as_series=False)` — pyarrow's
    `to_pydict()` produces value-identical Python scalars for the supported dtypes.
    """
    try:
        import pyarrow.parquet as pq
    except ImportError as e:
        raise ImportError(
            "Reading Parquet without Polars needs pyarrow: pip install goldencheck[parquet]"
        ) from e
    return pq.read_table(str(path)).to_pydict()


def _read_excel_columns(path: Path) -> dict[str, list]:
    """Read an Excel file into a dict[str, list] without Polars.

    Byte-identical to `pl.read_excel(path, engine="openpyxl").to_dict(as_series=False)`.
    Reads the FIRST worksheet (`wb.worksheets[0]`, not `wb.active` -- the saved-selected
    sheet, which can differ) and resolves formulas to their cached values (`data_only=True`).
    Raw openpyxl cell values are then coerced per-column to reproduce Polars' dtype
    inference, pinned empirically against `pl.read_excel`.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        return {}
    raw: dict[str, list] = {h: [] for h in header}
    for row in rows:
        for i, h in enumerate(header):
            raw[h].append(row[i] if i < len(row) else None)
    wb.close()
    return {h: _coerce_column(vals) for h, vals in raw.items()}


def _coerce_column(vals: list) -> list:
    """Reproduce pl.read_excel(engine='openpyxl')'s per-column coercion from raw openpyxl
    cell values. Empirically pinned against `pl.read_excel` (see P4a Task 2 investigation):

    - All-null / homogeneous columns pass through unchanged (ints, floats, strs, dates all
      already come back from openpyxl as the same Python types Polars reports).
    - Any `str` present in the column -> every non-null value stringifies. Bools stringify
      as lowercase `"true"`/`"false"` (NOT Python's `str(True) == "True"`); everything else
      uses plain `str(v)`.
    - No `str` present, column is pure `bool` -> stays bool (Polars Boolean dtype).
    - No `str` present, `float` appears anywhere (alone, or mixed with `int`/`bool`) ->
      every non-null value becomes `float` (bool True/False -> 1.0/0.0).
    - No `str`/`float` present, column mixes `bool` and `int` -> every non-null value
      becomes `int` (bool True/False -> 1/0).
    """
    non_null = [v for v in vals if v is not None]
    if not non_null:
        return list(vals)
    types = {type(v) for v in non_null}
    if str in types:

        def _stringify(v: object) -> str | None:
            if v is None:
                return None
            if isinstance(v, bool):
                return "true" if v else "false"
            return str(v)

        return [_stringify(v) for v in vals]
    if types <= {bool}:
        return list(vals)
    if float in types:
        return [None if v is None else float(v) for v in vals]
    if types <= {bool, int}:
        return [None if v is None else int(v) for v in vals]
    return list(vals)
